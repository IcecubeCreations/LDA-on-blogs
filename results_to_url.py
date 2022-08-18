import os
import sys
import openpyxl


def match_file_to_url(f, uf):
    for [url, filename] in uf:
        if filename == f:
            print(url + " = " + filename)
            return url
    return "None: " + f


def main():
    results_file_name = "topics_to_url.xlsx"
    if len(sys.argv) == 2 and sys.argv[1].endswith(".xlsx"):
        results_file_name = sys.argv[1]
    url_filename = []
    match_filename = []
    wb_pages = openpyxl.load_workbook("import_results.xlsx")
    wb_results = openpyxl.load_workbook("lda_result_names.xlsx")
    print(wb_pages.sheetnames)
    ws = wb_pages.active
    ws2 = wb_results.active
    for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
        url_filename.append([row[0], row[2].replace('fulltxt/', '')])
    
    for row in ws2.iter_rows(min_row=1, max_col=3, values_only=True):
        match_filename.append([row[0], row[1].replace('fulltxt/category_1_folder\\', ''), row[2]])

    print(len(url_filename))
    print(len(match_filename))

    wb_final = openpyxl.Workbook()
    ws3 = wb_final.active
    for index, [match, filename, score] in enumerate(sorted(match_filename)):
        ws3.cell(row = index +1, column = 1).value = match
        ws3.cell(row = index +1, column = 2).value = match_file_to_url(filename, url_filename)
        ws3.cell(row = index +1, column = 3).value = filename
        ws3.cell(row = index +1, column = 4).value = score 
    wb_final.save(results_file_name)

    url_filename.sort(key= lambda x: x[1])

    for [u, f] in url_filename:
        print(f)

if __name__ == "__main__":
    main()